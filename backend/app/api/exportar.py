# -*- coding: utf-8 -*-
"""Exportar presupuesto a Excel y PDF con logo del contratista."""
import json
import io
import logging
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from app.db import get_db, Usuario, Proyecto
from app.api.auth import usuario_actual
from app.services.calculo_presupuesto import calcular_totales

logger = logging.getLogger(__name__)
router = APIRouter()


class ExportRequest(BaseModel):
    proyecto_id: int


def _cargar(proyecto_id: int, user: Usuario, db: Session):
    p = db.query(Proyecto).filter(Proyecto.id == proyecto_id, Proyecto.user_id == user.id).first()
    if not p:
        raise HTTPException(404, "Proyecto no encontrado")
    items = json.loads(p.items_json or "[]")
    if not items:
        raise HTTPException(400, "El presupuesto esta vacio")
    aiu = json.loads(p.aiu_json or "{}")
    return p, items, aiu, calcular_totales(items, aiu)


@router.post("/excel")
def exportar_excel(req: ExportRequest, user: Usuario = Depends(usuario_actual), db: Session = Depends(get_db)):
    p, items, aiu, tot = _cargar(req.proyecto_id, user, db)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Presupuesto"

        azul = PatternFill("solid", fgColor="1C3A5E")
        gris = PatternFill("solid", fgColor="F1F5F9")
        blanco_bold = Font(color="FFFFFF", bold=True, size=11)
        bold = Font(bold=True)
        thin = Border(bottom=Side(style="thin", color="E2E8F0"))

        # Encabezado
        ws["A1"] = user.empresa or user.nombre
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Cotizacion {p.numero or ''} — {p.nombre}"
        ws["A2"].font = Font(bold=True, size=11)
        ws["A3"] = f"Cliente: {p.cliente_nombre or '-'}  |  {p.direccion or ''}"
        ws["A4"] = f"Fecha: {p.actualizado.strftime('%d/%m/%Y') if p.actualizado else ''}  |  Contacto: {user.telefono or user.email}"

        fila = 6
        headers = ["Item", "Capitulo", "Descripcion", "Und", "Cantidad", "Vr. Unitario", "Vr. Total"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=fila, column=col, value=h)
            c.fill = azul; c.font = blanco_bold
            c.alignment = Alignment(horizontal="center")
        fila += 1

        _incid = {c["capitulo"]: c["pct"] for c in tot.get("capitulos", [])}
        cap_actual = None
        n = 0
        num_cap = 0          # numeracion jerarquica del oficio: 1, 1.1, 1.2, 2, 2.1...
        for it in items:
            cap = it.get("capitulo", "OTROS")
            if cap != cap_actual:
                cap_actual = cap
                num_cap += 1
                n = 0
                pct_cap = _incid.get(cap)
                c = ws.cell(row=fila, column=1,
                            value=f"{num_cap}  {cap}" + (f"  ·  incidencia {pct_cap}%" if pct_cap is not None else ""))
                ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=7)
                c.fill = gris; c.font = bold
                fila += 1
            n += 1
            total_item = round(it["cantidad"] * it["precio_unitario"])
            vals = [f"{num_cap}.{n}", cap, it["descripcion"], it["unidad"], it["cantidad"], it["precio_unitario"], total_item]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=fila, column=col, value=v)
                c.border = thin
                if col in (6, 7): c.number_format = '"$"#,##0'
            fila += 1

        # Totales
        fila += 1
        def _linea(label, valor, es_total=False):
            nonlocal fila
            c1 = ws.cell(row=fila, column=6, value=label)
            c2 = ws.cell(row=fila, column=7, value=valor)
            c2.number_format = '"$"#,##0'
            if es_total:
                c1.font = bold; c2.font = bold
                c1.fill = gris; c2.fill = gris
            fila += 1

        if tot.get("descuento_valor"):
            _linea("Valor de lista", tot["subtotal_lista"])
            _linea(f"Descuento comercial {tot['descuento_pct']}%", -tot["descuento_valor"])
        _linea("Costo directo", tot["subtotal_directo"])
        if tot["aiu_total"] > 0:
            _linea(f"Administracion {tot['admin_pct']}%", tot["admin_valor"])
            _linea(f"Imprevistos {tot['imprevistos_pct']}%", tot["imprevistos_valor"])
            _linea(f"Utilidad {tot['utilidad_pct']}%", tot["utilidad_valor"])
        _linea("IVA", tot["iva"])
        _linea("TOTAL", tot["total"], es_total=True)
        c1 = ws.cell(row=fila, column=6, value="VALOR TOTAL OBRA")
        c2 = ws.cell(row=fila, column=7, value=tot["total"])
        c2.number_format = '"$"#,##0'
        c1.font = c2.font = Font(bold=True, size=12, color="1C3A5E")
        fila += 1

        ws.cell(row=fila + 1, column=1, value=tot["regimen_iva"]).font = Font(size=9, italic=True, color="64748B")
        fila += 2

        # Condiciones estandar del contratista
        condiciones = getattr(user, "condiciones", "") or ""
        if condiciones:
            ws.cell(row=fila + 1, column=1, value="CONDICIONES:").font = Font(bold=True, size=9)
            fila += 1
            for linea in condiciones.split("\n")[:15]:
                ws.cell(row=fila + 1, column=1, value=linea[:150]).font = Font(size=8, color="475569")
                fila += 1

        # Marca de agua plan gratis (motor de crecimiento)
        if user.plan == "gratis":
            fila += 1
            c = ws.cell(row=fila + 1, column=1, value="Hecho con PresupuestarCO — presupuestos profesionales en minutos · presupuestar.co")
            c.font = Font(size=8, italic=True, color="94A3B8")

        # ELABORO: la firma del formato del oficio
        fila += 2
        ws.cell(row=fila, column=1, value="ELABORÓ:").font = bold
        ws.cell(row=fila, column=2, value=user.nombre or "")
        ws.cell(row=fila + 1, column=2, value=user.empresa or "").font = Font(size=9, color="64748B")
        ws.cell(row=fila + 2, column=2, value="_______________________  (firma)").font = Font(size=9, color="94A3B8")
        fila += 3

        # Anchos
        for col, w in zip("ABCDEFG", [8, 22, 55, 8, 10, 14, 16]):
            ws.column_dimensions[col].width = w

        # ══ HOJA 2: MEMORIA DE CANTIDADES — cada cantidad, justificada ══════
        # El formato del oficio: ACTIVIDAD | UND | CANT | LARGO | ANCHO | ALTO | TOTAL.
        # Sale de it["calc"] (la calculadora 📐) — SIEMPRE del dato vivo.
        ws2 = wb.create_sheet("Memoria de cantidades")
        ws2["A1"] = "MEMORIA DE CANTIDADES"
        ws2["A1"].font = Font(bold=True, size=13, color="1C3A5E")
        ws2["A2"] = f"{p.nombre} — justificación de cada cantidad del presupuesto"
        ws2["A2"].font = Font(size=9, color="64748B")
        f2 = 4
        heads2 = ["Actividad", "Und", "# Elem.", "Largo (m)", "Ancho (m)", "Alto (m)", "Cantidad total"]
        for col, h in enumerate(heads2, 1):
            c = ws2.cell(row=f2, column=col, value=h)
            c.fill = azul; c.font = blanco_bold
        f2 += 1
        cap_actual2 = None
        con_memoria = 0
        for it in items:
            cap = it.get("capitulo", "OTROS")
            if cap != cap_actual2:
                cap_actual2 = cap
                c = ws2.cell(row=f2, column=1, value=cap)
                ws2.merge_cells(start_row=f2, start_column=1, end_row=f2, end_column=7)
                c.fill = gris; c.font = bold
                f2 += 1
            calc = it.get("calc") or {}
            tiene = bool(calc.get("largo") or calc.get("n"))
            if tiene:
                con_memoria += 1
            vals2 = [it["descripcion"][:90], it.get("unidad", ""),
                     calc.get("n") or (1 if tiene else ""),
                     calc.get("largo") or "", calc.get("ancho") or "", calc.get("alto") or "",
                     it["cantidad"]]
            for col, v in enumerate(vals2, 1):
                c = ws2.cell(row=f2, column=col, value=v)
                c.border = thin
            if not tiene:
                ws2.cell(row=f2, column=3, value="—").font = Font(color="94A3B8")
                ws2.cell(row=f2 + 0, column=4, value="cantidad directa").font = Font(size=8, italic=True, color="94A3B8")
            f2 += 1
        ws2.cell(row=f2 + 1, column=1,
                 value=f"{con_memoria} de {len(items)} cantidades con memoria dimensional (calculadora 📐). "
                       f"Las demás son cantidades directas del presupuestador.").font = Font(size=8, italic=True, color="64748B")
        for col, w in zip("ABCDEFG", [55, 8, 9, 10, 10, 10, 14]):
            ws2.column_dimensions[col].width = w

        # ══ HOJA 3: MATERIALES — la explosion de insumos, en el mismo libro ══
        ws3 = wb.create_sheet("Materiales")
        ws3["A1"] = "CANTIDAD DE MATERIALES"
        ws3["A1"].font = Font(bold=True, size=13, color="1C3A5E")
        from app.services.explosion_service import explosion_de_insumos
        rexp = explosion_de_insumos(items, _resolutor_de_desgloses(user, db),
                                    _precios_negociados(user, db))
        ws3["A2"] = (f"{rexp['items_explotados']} actividades explotadas · desperdicio incluido · "
                     f"precios: tus proveedores primero, luego el APU")
        ws3["A2"].font = Font(size=9, color="64748B")
        f3 = 4
        for col, h in enumerate(["Material", "Und", "Cantidad", "Precio", "Fuente", "Valor total"], 1):
            c = ws3.cell(row=f3, column=col, value=h)
            c.fill = azul; c.font = blanco_bold
        f3 += 1
        for i3 in rexp["insumos"]:
            vals3 = [i3["nombre"], i3["unidad"], i3["cantidad"], i3["precio"] or "",
                     (f"🏪 {i3['proveedor']} ({i3['fecha_precio']})" if i3["fuente_precio"] == "mi_proveedor"
                      else ("precio del APU" if i3["fuente_precio"] == "apu" else "")),
                     i3["subtotal"] or ""]
            for col, v in enumerate(vals3, 1):
                c = ws3.cell(row=f3, column=col, value=v)
                c.border = thin
                if col in (4, 6) and v != "":
                    c.number_format = '"$"#,##0'
            f3 += 1
        c1 = ws3.cell(row=f3, column=5, value="TOTAL MATERIALES")
        c2 = ws3.cell(row=f3, column=6, value=rexp["total_estimado"])
        c2.number_format = '"$"#,##0'
        c1.font = c2.font = bold
        c1.fill = c2.fill = gris
        if rexp["items_sin_desglose"]:
            ws3.cell(row=f3 + 2, column=1,
                     value=f"{rexp['items_sin_desglose']} actividad(es) sin desglose no entran "
                           f"(construyelas por insumos 🧱): " + "; ".join(rexp["sin_desglose"][:6])).font = \
                Font(size=8, italic=True, color="B45309")
        for col, w in zip("ABCDEF", [45, 8, 12, 13, 30, 15]):
            ws3.column_dimensions[col].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="presupuesto_{p.id}.xlsx"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Excel: {e}", exc_info=True)
        raise HTTPException(500, f"Error generando Excel: {e}")


@router.post("/pdf")
def exportar_pdf(req: ExportRequest, user: Usuario = Depends(usuario_actual), db: Session = Depends(get_db)):
    p, items, aiu, tot = _cargar(req.proyecto_id, user, db)
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer, Image as RLImage)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=1.5*cm, bottomMargin=1.5*cm,
                                leftMargin=1.5*cm, rightMargin=1.5*cm)
        styles = getSampleStyleSheet()
        story = []

        # Logo si existe
        if user.logo_b64:
            try:
                import base64
                b64 = user.logo_b64.split(",")[-1]
                logo_bytes = base64.b64decode(b64)
                img = RLImage(io.BytesIO(logo_bytes), width=3*cm, height=3*cm, kind='proportional')
                story.append(img)
                story.append(Spacer(1, 8))
            except Exception:
                pass

        titulo = ParagraphStyle("t", parent=styles["Title"], fontSize=16, textColor=colors.HexColor("#1C3A5E"), alignment=0)
        story.append(Paragraph(user.empresa or user.nombre, titulo))
        story.append(Paragraph(f"<b>Cotizacion No.:</b> {p.numero or '-'} &nbsp;&nbsp; <b>Proyecto:</b> {p.nombre}", styles["Normal"]))
        story.append(Paragraph(f"<b>Cliente:</b> {p.cliente_nombre or '-'} | {p.direccion or ''}", styles["Normal"]))
        story.append(Paragraph(
            f"<b>Fecha:</b> {p.actualizado.strftime('%d/%m/%Y') if p.actualizado else ''} | "
            f"<b>Contacto:</b> {user.telefono or user.email}", styles["Normal"]))
        story.append(Spacer(1, 14))

        # Tabla de items agrupada por capitulo
        data = [["#", "Descripcion", "Und", "Cant", "Vr. Unit", "Vr. Total"]]
        estilos_extra = []
        fila_idx = 1
        _incid = {c["capitulo"]: c["pct"] for c in tot.get("capitulos", [])}
        cap_actual = None
        n = 0
        for it in items:
            cap = it.get("capitulo", "OTROS")
            if cap != cap_actual:
                cap_actual = cap
                _pc = _incid.get(cap)
                data.append([f"{cap}  ·  incidencia {_pc}%" if _pc is not None else cap, "", "", "", "", ""])
                estilos_extra.append(("SPAN", (0, fila_idx), (-1, fila_idx)))
                estilos_extra.append(("BACKGROUND", (0, fila_idx), (-1, fila_idx), colors.HexColor("#F1F5F9")))
                estilos_extra.append(("FONTNAME", (0, fila_idx), (-1, fila_idx), "Helvetica-Bold"))
                fila_idx += 1
            n += 1
            total_item = round(it["cantidad"] * it["precio_unitario"])
            desc = it["descripcion"][:80]
            data.append([str(n), Paragraph(desc, styles["BodyText"]), it["unidad"],
                         f"{it['cantidad']:g}", f"${it['precio_unitario']:,.0f}", f"${total_item:,.0f}"])
            fila_idx += 1

        tabla = Table(data, colWidths=[1*cm, 9.5*cm, 1.3*cm, 1.6*cm, 2.5*cm, 2.6*cm], repeatRows=1)
        tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1C3A5E")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
            ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ] + estilos_extra))
        story.append(tabla)
        story.append(Spacer(1, 14))

        # Resumen
        resumen = []
        if tot.get("descuento_valor"):
            resumen += [["Valor de lista", f"${tot['subtotal_lista']:,.0f}"],
                        [f"Descuento comercial {tot['descuento_pct']}%", f"-${tot['descuento_valor']:,.0f}"]]
        resumen += [["Costo directo", f"${tot['subtotal_directo']:,.0f}"]]
        if tot["aiu_total"] > 0:
            resumen += [
                [f"Administracion {tot['admin_pct']}%", f"${tot['admin_valor']:,.0f}"],
                [f"Imprevistos {tot['imprevistos_pct']}%", f"${tot['imprevistos_valor']:,.0f}"],
                [f"Utilidad {tot['utilidad_pct']}%", f"${tot['utilidad_valor']:,.0f}"],
            ]
        resumen += [["IVA", f"${tot['iva']:,.0f}"], ["TOTAL", f"${tot['total']:,.0f}"]]
        tr = Table(resumen, colWidths=[13*cm, 5.5*cm])
        tr.setStyle(TableStyle([
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F5F9")),
            ("LINEABOVE", (0, -1), (-1, -1), 1, colors.HexColor("#1C3A5E")),
        ]))
        story.append(tr)
        story.append(Spacer(1, 6))
        story.append(Paragraph(f"<i>{tot['regimen_iva']}</i>",
                               ParagraphStyle("n", parent=styles["Normal"], fontSize=7, textColor=colors.HexColor("#64748B"))))
        if p.notas:
            story.append(Spacer(1, 10))
            story.append(Paragraph(f"<b>Notas:</b> {p.notas[:800]}", styles["Normal"]))

        # Condiciones estandar del contratista
        condiciones = getattr(user, "condiciones", "") or ""
        if condiciones:
            story.append(Spacer(1, 10))
            story.append(Paragraph("<b>Condiciones:</b>", styles["Normal"]))
            cond_style = ParagraphStyle("cond", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#475569"))
            for linea in condiciones.split("\n")[:15]:
                if linea.strip():
                    story.append(Paragraph(linea.strip()[:200], cond_style))

        # Marca de agua plan gratis
        if user.plan == "gratis":
            story.append(Spacer(1, 16))
            marca = ParagraphStyle("marca", parent=styles["Normal"], fontSize=7,
                                   textColor=colors.HexColor("#94A3B8"), alignment=1)
            story.append(Paragraph("Hecho con <b>PresupuestarCO</b> — presupuestos profesionales en minutos", marca))

        doc.build(story)
        buf.seek(0)
        return StreamingResponse(
            buf, media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="presupuesto_{p.id}.pdf"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"PDF: {e}", exc_info=True)
        raise HTTPException(500, f"Error generando PDF: {e}")


# ═══════════════════════════════════════════════════════════════════════════
# F1 — ANALISIS DE PRECIOS UNITARIOS (PDF formato propuesta)
# El anexo que toda propuesta estatal exige — y que profesionaliza la privada.
# Un APU por item: si el item nacio del CONSTRUCTOR (apu_id -> desglose_json),
# imprime el analisis completo (materiales+desperdicio, MO, herramienta);
# si viene de la base 2026, imprime la referencia con su factor de ciudad.
# ═══════════════════════════════════════════════════════════════════════════

def _resolutor_de_desgloses(user: Usuario, db: Session):
    """Mapa de desgloses del usuario + resolutor triple (apu_id / MIO-{id} / codigo).
    Compartido por el anexo de APUs y la explosion de insumos."""
    from app.db import ApuUsuario
    por_id, por_codigo = {}, {}
    for a in db.query(ApuUsuario).filter(ApuUsuario.user_id == user.id,
                                         ApuUsuario.desglose_json != "").all():
        try:
            dj = json.loads(a.desglose_json)
        except Exception:
            continue
        if not dj or not dj.get("insumos"):
            continue
        por_id[a.id] = dj
        if (a.codigo or "").strip():
            por_codigo[a.codigo.strip()] = dj

    def _desglose_de(it):
        if it.get("apu_id") and it["apu_id"] in por_id:
            return por_id[it["apu_id"]]
        cod = (it.get("codigo") or "").strip()
        if cod.startswith("MIO-") and cod[4:].isdigit() and int(cod[4:]) in por_id:
            return por_id[int(cod[4:])]
        return por_codigo.get(cod)
    return _desglose_de


@router.post("/apus-pdf")
def exportar_apus_pdf(req: ExportRequest, user: Usuario = Depends(usuario_actual), db: Session = Depends(get_db)):
    p, items, aiu, tot = _cargar(req.proyecto_id, user, db)
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer, PageBreak)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from app.db import ApuUsuario
        from app.services.apu_service import FACTORES_REGION

        es_publico = (p.sector or "privado") == "publico"
        region = FACTORES_REGION.get(p.region or "bogota", FACTORES_REGION["bogota"])

        # Desgloses del usuario: resolutor triple compartido (apu_id / MIO / codigo)
        _desglose_de = _resolutor_de_desgloses(user, db)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=1.4*cm, bottomMargin=1.4*cm,
                                leftMargin=1.5*cm, rightMargin=1.5*cm)
        styles = getSampleStyleSheet()
        NAVY, GRIS, AMBAR = colors.HexColor("#1C3A5E"), colors.HexColor("#64748B"), colors.HexColor("#B45309")
        st_t = ParagraphStyle("t", parent=styles["Title"], fontSize=15, textColor=NAVY, alignment=1)
        st_sub = ParagraphStyle("s", parent=styles["Normal"], fontSize=9, textColor=GRIS, alignment=1)
        st_h = ParagraphStyle("h", parent=styles["Normal"], fontSize=10, textColor=NAVY, spaceBefore=4)
        st_n = ParagraphStyle("n", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#1A1A1A"))
        st_nota = ParagraphStyle("no", parent=styles["Normal"], fontSize=7.5, textColor=GRIS)
        COP = lambda v: f"${round(v or 0):,.0f}".replace(",", ".")
        story = []

        # Logo del contratista (R4): el documento de venta mas fuerte, con marca propia
        if user.logo_b64:
            try:
                import base64
                from reportlab.platypus import Image as RLImage
                b64 = user.logo_b64.split(",")[-1]
                img = RLImage(io.BytesIO(base64.b64decode(b64)), width=2.6*cm, height=2.6*cm, kind='proportional')
                img.hAlign = "CENTER"
                story.append(img)
                story.append(Spacer(1, 6))
            except Exception:
                pass

        # ── Portada ──
        story.append(Paragraph("ANÁLISIS DE PRECIOS UNITARIOS", st_t))
        story.append(Paragraph("Anexo de la propuesta económica" if es_publico
                               else "Soporte técnico del presupuesto", st_sub))
        story.append(Spacer(1, 10))
        quien = [("Entidad contratante", p.entidad_nombre or "-"),
                 ("Contrato No.", p.contrato_numero or "-"),
                 ("Supervisor", p.supervisor_nombre or "-")] if es_publico else \
                [("Cliente", p.cliente_nombre or "-"), ("Dirección", p.direccion or "-")]
        cab = [["Proyecto", p.nombre], *quien,
               ["Proponente" if es_publico else "Contratista", user.empresa or user.nombre],
               ["Ciudad de referencia", f"{region['nombre']} (factor {region['factor']})"],
               ["AIU aplicado", f"A {tot['admin_pct']}% · I {tot['imprevistos_pct']}% · U {tot['utilidad_pct']}%"
                                if aiu.get("aplicar", True) else "Sin AIU"],
               ["Ítems analizados", str(len(items))],
               ["Fecha", p.actualizado.strftime("%d/%m/%Y") if p.actualizado else ""]]
        t = Table([[Paragraph(f"<b>{k}</b>", st_n), Paragraph(str(v)[:120], st_n)] for k, v in cab],
                  colWidths=[4.5*cm, 12*cm])
        t.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
                               ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F1F5F9")),
                               ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]))
        story.append(t)
        story.append(Spacer(1, 6))
        story.append(Paragraph(
            "Cada análisis desglosa materiales (con % de desperdicio), mano de obra y herramienta menor. "
            "Los ítems tomados de la base de referencia 2026 se presentan con su precio unitario ajustado "
            "por el factor de la ciudad. El desperdicio cubre cortes, mermas y manipulación — práctica "
            "estándar de presupuestación.", st_nota))
        story.append(PageBreak())

        # ── Un APU por item ──
        for idx, it in enumerate(items, 1):
            cant = float(it.get("cantidad") or 0)
            pu = round(float(it.get("precio_unitario") or 0))
            story.append(Paragraph(f"APU {idx:02d} — {(it.get('codigo') or '').strip() or 's/c'}", st_h))
            story.append(Paragraph(f"<b>{(it.get('descripcion') or 'Ítem')[:180]}</b>", st_n))
            story.append(Paragraph(
                f"Capítulo: {(it.get('capitulo') or 'OTROS')[:60]} · Unidad: {it.get('unidad') or 'un'} · "
                f"Cantidad: {cant:g} · <b>Vr. unitario: {COP(pu)}</b> · Vr. total: {COP(cant*pu)}", st_n))
            pl = float(it.get("precio_lista") or 0)
            if pl > pu:
                story.append(Paragraph(
                    f"Precio de lista {COP(pl)} — descuento comercial pactado del "
                    f"{round((pl-pu)*100/pl, 1)}% ya reflejado en el unitario.", st_nota))
            story.append(Spacer(1, 3))

            d = _desglose_de(it)
            if d and d.get("insumos"):
                filas = [["1. MATERIALES", "Und", "Cant.", "Desp.%", "Vr. unit.", "Parcial"]]
                for ins in d["insumos"][:40]:
                    filas.append([str(ins.get("nombre") or "")[:48], str(ins.get("unidad") or "")[:6],
                                  f"{float(ins.get('cantidad') or 0):g}",
                                  f"{float(ins.get('desperdicio_pct') or 0):g}",
                                  COP(ins.get("precio")), COP(ins.get("parcial"))])
                filas.append(["Subtotal materiales", "", "", "", "", COP(d.get("materiales"))])
                filas.append(["2. MANO DE OBRA (por unidad)", "", "", "", "", COP(d.get("mano_obra"))])
                filas.append([f"3. HERRAMIENTA MENOR ({float(d.get('herramienta_pct') or 0):g}% de MO)",
                              "", "", "", "", COP(d.get("herramienta"))])
                filas.append(["4. TRANSPORTE (por unidad)", "", "", "", "", COP(d.get("transporte") or 0)])
                filas.append(["PRECIO UNITARIO ANALIZADO", "", "", "", "", COP(d.get("precio_unitario"))])
                ta = Table(filas, colWidths=[7.2*cm, 1.4*cm, 1.6*cm, 1.6*cm, 2.5*cm, 2.7*cm])
                ta.setStyle(TableStyle([
                    ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                    ("BACKGROUND", (0, 0), (-1, 0), NAVY), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
                    ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
                    ("BACKGROUND", (0, -5), (-1, -1), colors.HexColor("#F1F5F9")),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("TOPPADDING", (0, 0), (-1, -1), 2), ("BOTTOMPADDING", (0, 0), (-1, -1), 2)]))
                story.append(ta)
                analizado = round(float(d.get("precio_unitario") or 0))
                if analizado and analizado != pu:
                    story.append(Paragraph(
                        f"Nota: el precio pactado ({COP(pu)}) difiere del analizado ({COP(analizado)}) — "
                        f"ajuste comercial del proponente.",
                        ParagraphStyle("w", parent=st_nota, textColor=AMBAR)))
            elif (it.get("codigo") or "").strip() and not str(it.get("codigo")).startswith("MIO"):
                story.append(Paragraph(
                    f"Precio unitario de <b>referencia — Base APU 2026</b> (código {it.get('codigo')}), "
                    f"ajustado a {region['nombre']} con factor {region['factor']}. Incluye materiales, "
                    f"mano de obra, herramienta y desperdicios propios de la actividad según la base.", st_nota))
            else:
                story.append(Paragraph(
                    "APU propio a precio global del proponente (sin desglose registrado). "
                    "Constrúyelo por insumos en la plataforma para anexar el análisis completo.", st_nota))
            story.append(Spacer(1, 10))

        if user.plan == "gratis":
            story.append(Paragraph("Hecho con <b>PresupuestarCO</b> — del APU al acta, todo conectado",
                                   ParagraphStyle("m", parent=st_nota, alignment=1)))
        doc.build(story)
        buf.seek(0)
        return StreamingResponse(
            buf, media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="apus_{p.id}.pdf"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"APUs PDF: {e}", exc_info=True)
        raise HTTPException(500, f"Error generando el anexo de APUs: {e}")


# ═══════════════════════════════════════════════════════════════════════════
# F2 — EXPLOSION DE INSUMOS: la lista de compras de la obra completa
# Consolida los insumos de todos los APUs con desglose (con desperdicio) y los
# cruza con los precios negociados de MIS proveedores. JSON para la UI + Excel.
# ═══════════════════════════════════════════════════════════════════════════

def _precios_negociados(user: Usuario, db: Session) -> dict:
    """Mejor precio por insumo entre MIS proveedores: {nombre_lower: {...}}."""
    from app.db import PrecioProveedor, Proveedor
    filas = (db.query(PrecioProveedor, Proveedor.nombre)
             .join(Proveedor, Proveedor.id == PrecioProveedor.proveedor_id)
             .filter(PrecioProveedor.user_id == user.id,
                     PrecioProveedor.precio > 0).all())
    mejor = {}
    for pp, prov_nombre in filas:
        k = (pp.insumo or "").strip().lower()
        if not k:
            continue
        if k not in mejor or pp.precio < mejor[k]["precio"]:
            mejor[k] = {"precio": pp.precio, "proveedor": prov_nombre,
                        "fecha": pp.capturado.strftime("%d/%m/%Y") if pp.capturado else ""}
    return mejor


@router.post("/explosion")
def explosion(req: ExportRequest, user: Usuario = Depends(usuario_actual), db: Session = Depends(get_db)):
    """La lista de materiales consolidada, con el parte honesto de cobertura."""
    p, items, aiu, tot = _cargar(req.proyecto_id, user, db)
    from app.services.explosion_service import explosion_de_insumos
    resultado = explosion_de_insumos(items, _resolutor_de_desgloses(user, db),
                                     _precios_negociados(user, db))
    resultado["proyecto"] = p.nombre
    return resultado


@router.post("/explosion-excel")
def explosion_excel(req: ExportRequest, user: Usuario = Depends(usuario_actual), db: Session = Depends(get_db)):
    """La misma lista, lista para imprimir y llevar a la ferreteria."""
    p, items, aiu, tot = _cargar(req.proyecto_id, user, db)
    from app.services.explosion_service import explosion_de_insumos
    r = explosion_de_insumos(items, _resolutor_de_desgloses(user, db),
                             _precios_negociados(user, db))
    if not r["insumos"]:
        raise HTTPException(400, "Ningun item tiene desglose de insumos todavia — "
                                 "construye tus APUs por insumos para generar la lista de compras")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Lista de materiales"
        azul = PatternFill("solid", fgColor="1C3A5E")
        gris = PatternFill("solid", fgColor="F1F5F9")
        thin = Border(bottom=Side(style="thin", color="E2E8F0"))

        ws["A1"] = user.empresa or user.nombre
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"LISTA DE MATERIALES — {p.nombre}"
        ws["A2"].font = Font(bold=True, size=11)
        ws["A3"] = (f"{r['items_explotados']} actividades explotadas · "
                    f"{r['n_insumos']} insumos consolidados (desperdicio incluido) · "
                    f"Fecha: {p.actualizado.strftime('%d/%m/%Y') if p.actualizado else ''}")
        ws["A3"].font = Font(size=9, color="64748B")

        fila = 5
        headers = ["Insumo", "Und", "Cantidad total", "Mejor precio", "Proveedor", "Precio al", "Subtotal"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=fila, column=col, value=h)
            c.fill = azul
            c.font = Font(color="FFFFFF", bold=True, size=10)
        fila += 1
        for i in r["insumos"]:
            vals = [i["nombre"], i["unidad"], i["cantidad"],
                    i["precio"] or "", i["proveedor"] or ("precio del APU" if i["fuente_precio"] == "apu" else ""),
                    i["fecha_precio"] or "", i["subtotal"] or ""]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=fila, column=col, value=v)
                c.border = thin
                if col in (4, 7) and v != "":
                    c.number_format = '"$"#,##0'
            fila += 1
        c1 = ws.cell(row=fila, column=6, value="TOTAL MATERIALES ESTIMADO")
        c2 = ws.cell(row=fila, column=7, value=r["total_estimado"])
        c2.number_format = '"$"#,##0'
        c1.font = c2.font = Font(bold=True)
        c1.fill = c2.fill = gris
        if r["items_sin_desglose"]:
            fila += 2
            ws.cell(row=fila, column=1,
                    value=f"Nota: {r['items_sin_desglose']} actividad(es) sin desglose no entran a la lista "
                          f"(construyelas por insumos): " + "; ".join(r["sin_desglose"])).font = \
                Font(size=8, italic=True, color="B45309")
        anchos = [42, 8, 14, 14, 24, 12, 14]
        from openpyxl.utils import get_column_letter
        for i2, a in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(i2)].width = a

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="materiales_{p.id}.xlsx"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Explosion Excel: {e}", exc_info=True)
        raise HTTPException(500, f"Error generando la lista de materiales: {e}")


# ═══════════════════════════════════════════════════════════════════════════
# CRONOGRAMA DE OBRA — PDF con el diagrama de Gantt (solo sector publico)
# ═══════════════════════════════════════════════════════════════════════════

@router.post("/cronograma-pdf")
def exportar_cronograma_pdf(req: ExportRequest, user: Usuario = Depends(usuario_actual),
                            db: Session = Depends(get_db)):
    p = db.query(Proyecto).filter(Proyecto.id == req.proyecto_id, Proyecto.user_id == user.id).first()
    if not p:
        raise HTTPException(404, "Proyecto no encontrado")
    if (p.sector or "privado") != "publico":
        raise HTTPException(400, "El cronograma es una herramienta exclusiva de obra publica")

    from app.services.cronograma_service import resolver_cronograma, duracion_total_semanas, CronogramaError
    data = json.loads(p.cronograma_json or "{}")
    filas_crudas = data.get("filas", [])
    if not filas_crudas:
        raise HTTPException(400, "El cronograma esta vacio — agrega filas antes de exportar")
    try:
        filas = resolver_cronograma(filas_crudas)
    except CronogramaError as e:
        raise HTTPException(400, f"El cronograma tiene un error: {e} — corrigelo antes de exportar")

    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        contrato = json.loads(p.contrato_json or "{}")
        import math
        total_semanas = max(1, math.ceil(duracion_total_semanas(filas)))

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(letter), topMargin=1.2*cm, bottomMargin=1.2*cm,
                                leftMargin=1.2*cm, rightMargin=1.2*cm)
        styles = getSampleStyleSheet()
        story = []

        titulo = ParagraphStyle("t", parent=styles["Title"], fontSize=15, textColor=colors.HexColor("#1C3A5E"), alignment=0)
        story.append(Paragraph(f"Cronograma de obra — {p.nombre}", titulo))
        sub = f"<b>Entidad:</b> {p.entidad_nombre or '-'} &nbsp;&nbsp; <b>Contrato:</b> {p.contrato_numero or '-'}"
        if contrato.get("fecha_inicio"):
            sub += f" &nbsp;&nbsp; <b>Inicio:</b> {contrato['fecha_inicio']}"
        story.append(Paragraph(sub, styles["Normal"]))
        story.append(Spacer(1, 12))

        # Ancho disponible: columna de nombre fija, el resto repartido entre las semanas
        ancho_nombre = 5.5 * cm
        ancho_disponible = landscape(letter)[0] - 2.4 * cm - ancho_nombre
        # Sin ancho minimo: para contratos largos (>~57 semanas con el minimo
        # anterior de 0.35cm) la tabla se desbordaba de la pagina. Mejor
        # columnas finas y legibles-por-agrupacion que una tabla rota.
        ancho_semana = ancho_disponible / total_semanas

        # Con muchas semanas los numeros se aplastan y quedan ilegibles (verificado
        # generando el PDF real con 61 columnas) — a partir de 20 semanas, mostrar
        # el numero solo cada tantas columnas, dejando el resto en blanco.
        intervalo = max(1, round(total_semanas / 20))
        encabezado = ["Actividad"] + [f"S{i+1}" if (i % intervalo == 0 or i == total_semanas - 1) else "" for i in range(total_semanas)]
        filas_tabla = [encabezado]
        estilos_barras = []
        for idx, f in enumerate(filas, start=1):
            fila_render = [Paragraph(f["nombre"][:60], styles["BodyText"])] + [""] * total_semanas
            filas_tabla.append(fila_render)
            ini = int(f["semana_inicio_efectiva"])
            fin = max(ini + 1, int(round(f["semana_fin_efectiva"])))
            for sem in range(ini, min(fin, total_semanas)):
                estilos_barras.append(("BACKGROUND", (sem + 1, idx), (sem + 1, idx), colors.HexColor("#0284C7")))

        tabla = Table(filas_tabla, colWidths=[ancho_nombre] + [ancho_semana] * total_semanas, repeatRows=1)
        tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1C3A5E")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 6.5),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ] + estilos_barras))
        story.append(tabla)
        story.append(Spacer(1, 14))

        from app.db import Avance
        ultimo = (db.query(Avance).filter(Avance.proyecto_id == p.id).order_by(Avance.creado.desc()).first())
        if ultimo:
            story.append(Paragraph(f"<b>Avance real cargado:</b> {ultimo.porcentaje}%", styles["Normal"]))

        story.append(Spacer(1, 10))
        marca = ParagraphStyle("marca", parent=styles["Normal"], fontSize=7,
                               textColor=colors.HexColor("#94A3B8"), alignment=1)
        story.append(Paragraph("Hecho con <b>PresupuestarCO</b> — presupuestos profesionales en minutos", marca))

        doc.build(story)
        buf.seek(0)
        return StreamingResponse(
            buf, media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="cronograma_{p.id}.pdf"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Cronograma PDF: {e}", exc_info=True)
        raise HTTPException(500, f"Error generando el PDF del cronograma: {e}")

