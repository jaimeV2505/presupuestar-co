"""Exportación profesional: Excel (APU detallado) y PDF."""
import io
from datetime import datetime
from app.models.schemas import PresupuestoResponse


def export_excel(p: PresupuestoResponse, nombre: str = "Proyecto") -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    NAVY = "1C3A5E"; STEEL = "2D6A9F"; LIGHT = "EBF2FA"; ALT = "F7FAFD"

    def fill(c): return PatternFill("solid", fgColor=c)
    def thin(): s = Side(style="thin", color="CCCCCC"); return Border(left=s, right=s, top=s, bottom=s)
    def hdr(sz=11, bold=True, col="FFFFFF"): return Font(name="Calibri", size=sz, bold=bold, color=col)
    def nrm(sz=9, bold=False, col="1A1A1A"): return Font(name="Calibri", size=sz, bold=bold, color=col)

    # ── Hoja 1: Presupuesto ───────────────────────────────────────────────────
    ws = wb.active; ws.title = "Presupuesto APU"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGH", [6,14,52,8,12,18,20,24]):
        ws.column_dimensions[get_column_letter(ord(col)-64)].width = w

    ws.merge_cells("A1:H1")
    c = ws["A1"]; c.value = f"PRESUPUESTO DE OBRA — {nombre.upper()}"
    c.font = hdr(14); c.fill = fill(NAVY); c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = f"Región: {p.region}  |  {p.fc_concreto}  |  Fecha: {datetime.now().strftime('%d/%m/%Y')}  |  Fuente: {p.fuente_precios}"
    c.font = hdr(9); c.fill = fill(STEEL); c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Cabeceras
    hdrs = ["#","Código","Ítem / Actividad","Und.","Cantidad","V. Unitario ($)","V. Total ($)","Rendimiento"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(4, ci, h); c.font = hdr(9); c.fill = fill(STEEL)
        c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = thin()
    ws.row_dimensions[4].height = 22

    row = 5
    by_cap = {}
    for item in p.items:
        by_cap.setdefault(item.capitulo, []).append(item)

    alt = False
    for cap, cap_items in by_cap.items():
        ws.merge_cells(f"A{row}:H{row}")
        c = ws.cell(row, 1, f"  {cap}"); c.font = hdr(9); c.fill = fill(STEEL)
        c.alignment = Alignment(horizontal="left"); ws.row_dimensions[row].height = 18; row += 1

        cap_total = 0
        for item in cap_items:
            fl = fill(ALT) if alt else PatternFill(); alt = not alt
            vals = [item.idx, item.codigo or "", item.nombre, item.unidad,
                    item.cantidad, item.precio_unitario, item.precio_total, item.rendimiento or ""]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row, ci, val); c.font = nrm(); c.border = thin(); c.fill = fl
                if ci in (5,6,7): c.alignment = Alignment(horizontal="right")
                if ci in (6,7): c.number_format = '"$"#,##0'
                elif ci == 3: c.alignment = Alignment(wrap_text=True)
                else: c.alignment = Alignment(horizontal="center")
            ws.row_dimensions[row].height = 15; cap_total += item.precio_total; row += 1

        for ci in range(1, 9):
            ws.cell(row, ci).fill = fill(LIGHT)
        ws.cell(row, 3, f"SUBTOTAL {cap}").font = nrm(bold=True, col=NAVY)
        ws.cell(row, 3).fill = fill(LIGHT)
        c = ws.cell(row, 7, cap_total); c.number_format = '"$"#,##0'
        c.font = nrm(bold=True, col=NAVY); c.fill = fill(LIGHT)
        c.alignment = Alignment(horizontal="right"); ws.row_dimensions[row].height = 16; row += 2

    for lbl, val, is_total in [
        ("SUBTOTAL DIRECTO", p.resumen.subtotal_directo, False),
        (f"AIU {p.resumen.porcentaje_aiu}%", p.resumen.aiu, False),
        ("BASE CON AIU", p.resumen.subtotal_con_aiu, False),
        (f"IVA {p.resumen.porcentaje_iva:.0f}%", p.resumen.iva, False),
        ("TOTAL PRESUPUESTO", p.resumen.total, True),
    ]:
        fc = NAVY if is_total else "D0E4F7"; fc_txt = "FFFFFF" if is_total else NAVY
        for ci in range(1, 9): ws.cell(row, ci).fill = fill(fc)
        ws.merge_cells(f"A{row}:F{row}")
        ws.cell(row, 1, f"  {lbl}").font = Font("Calibri", size=10 if is_total else 9,
                                                  bold=True, color=fc_txt)
        ws.cell(row, 1).alignment = Alignment(horizontal="right")
        c = ws.cell(row, 7, val); c.number_format = '"$"#,##0'
        c.font = Font("Calibri", size=10 if is_total else 9, bold=True, color=fc_txt)
        c.alignment = Alignment(horizontal="right")
        ws.row_dimensions[row].height = 18 if is_total else 15; row += 1

    if p.resumen.costo_m2:
        row += 1
        ws.cell(row, 3, f"Costo estimado por m²: ${p.resumen.costo_m2:,.0f} COP/m²").font = nrm(bold=True, col=STEEL)

    # ── Hoja 2: Resumen ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen Ejecutivo")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 38; ws2.column_dimensions["B"].width = 24
    ws2.merge_cells("A1:B1"); c = ws2["A1"]
    c.value = "RESUMEN EJECUTIVO"; c.font = hdr(13); c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center"); ws2.row_dimensions[1].height = 30

    for r, (k, v) in enumerate([
        ("Proyecto", nombre), ("Región", p.region), ("f'c concreto", p.fc_concreto),
        ("Fuente precios", p.fuente_precios), ("Fecha", p.fecha), ("", ""),
        ("Subtotal directo", f"${p.resumen.subtotal_directo:,.0f}"),
        (f"AIU {p.resumen.porcentaje_aiu}%", f"${p.resumen.aiu:,.0f}"),
        ("Base con AIU", f"${p.resumen.subtotal_con_aiu:,.0f}"),
        (f"IVA {p.resumen.porcentaje_iva:.0f}%", f"${p.resumen.iva:,.0f}"),
        ("TOTAL PRESUPUESTO", f"${p.resumen.total:,.0f}"),
    ], start=3):
        ws2.cell(r, 1, k).font = nrm(bold=True)
        c2 = ws2.cell(r, 2, v); c2.font = nrm()
        if k == "TOTAL PRESUPUESTO":
            ws2.cell(r, 1).fill = fill(NAVY); ws2.cell(r, 1).font = hdr(10, color=NAVY)
            c2.fill = fill(NAVY); c2.font = hdr(10)

    if p.resumen.costo_m2:
        ws2.cell(15, 1, "Costo/m²").font = nrm(bold=True)
        ws2.cell(15, 2, f"${p.resumen.costo_m2:,.0f} COP/m²").font = nrm()

    ws2.cell(17, 1, "NOTAS").font = Font("Calibri", size=10, bold=True, color=NAVY)
    for i, nota in enumerate(p.notas, 18):
        ws2.merge_cells(f"A{i}:B{i}")
        ws2.cell(i, 1, f"• {nota}").font = Font("Calibri", size=8, color="555555")
        ws2.row_dimensions[i].height = 16

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def export_pdf(p: PresupuestoResponse, nombre: str = "Proyecto") -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            title=f"Presupuesto — {nombre}")

    NAVY = colors.HexColor("#1C3A5E"); STEEL = colors.HexColor("#2D6A9F")
    LIGHT = colors.HexColor("#EBF2FA"); GRAY = colors.HexColor("#F7FAFD")
    BORDER = colors.HexColor("#CCCCCC")

    s_title = ParagraphStyle("t", fontSize=15, fontName="Helvetica-Bold",
                              textColor=colors.white, alignment=TA_CENTER)
    s_sub   = ParagraphStyle("s", fontSize=9,  fontName="Helvetica",
                              textColor=colors.white, alignment=TA_CENTER)
    s_cap   = ParagraphStyle("c", fontSize=8,  fontName="Helvetica-Bold",
                              textColor=colors.white, alignment=TA_LEFT)
    s_item  = ParagraphStyle("i", fontSize=7.5,fontName="Helvetica",
                              textColor=colors.HexColor("#1A1A1A"))
    s_note  = ParagraphStyle("n", fontSize=6.5,fontName="Helvetica",
                              textColor=colors.HexColor("#666666"), spaceAfter=2)

    story = []

    # Header
    hdr_data = [[Paragraph(f"PRESUPUESTO DE OBRA — {nombre.upper()}", s_title)]]
    t_hdr = Table(hdr_data, colWidths=["100%"])
    t_hdr.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),NAVY),
                                ("TOPPADDING",(0,0),(-1,-1),10),("BOTTOMPADDING",(0,0),(-1,-1),10)]))
    story.append(t_hdr)

    sub_data = [[Paragraph(
        f"Región: {p.region}  |  {p.fc_concreto}  |  Fecha: {datetime.now().strftime('%d/%m/%Y')}  |  {p.fuente_precios}", s_sub)]]
    t_sub = Table(sub_data, colWidths=["100%"])
    t_sub.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),STEEL),
                                ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]))
    story.append(t_sub); story.append(Spacer(1, 0.3*cm))

    # Tabla principal
    COL_W = [1*cm, 2.2*cm, 10.5*cm, 1.5*cm, 2*cm, 3*cm, 3*cm, 3.8*cm]
    table_data = [["#","Código","Ítem / Actividad","Und.","Cant.","V. Unitario ($)","V. Total ($)","Rendimiento"]]
    styles = [
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),8),
        ("BACKGROUND",(0,0),(-1,0),STEEL),("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),("ALIGN",(2,1),(2,-1),"LEFT"),
        ("ALIGN",(4,1),(6,-1),"RIGHT"),("GRID",(0,0),(-1,-1),0.3,BORDER),
        ("FONTNAME",(0,1),(-1,-1),"Helvetica"),("FONTSIZE",(0,1),(-1,-1),7),
        ("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2),
    ]
    row_i = 1
    by_cap = {}
    for item in p.items: by_cap.setdefault(item.capitulo, []).append(item)

    for cap, cap_items in by_cap.items():
        table_data.append([Paragraph(f"  {cap}", s_cap),"","","","","","",""])
        styles += [("BACKGROUND",(0,row_i),(-1,row_i),STEEL),("SPAN",(0,row_i),(-1,row_i))]
        row_i += 1
        cap_total = 0
        for item in cap_items:
            table_data.append([str(item.idx), item.codigo or "",
                                Paragraph(item.nombre, s_item),
                                item.unidad,
                                f"{item.cantidad:,.2f}",
                                f"${item.precio_unitario:,.0f}",
                                f"${item.precio_total:,.0f}",
                                item.rendimiento or ""])
            if row_i % 2 == 0: styles.append(("BACKGROUND",(0,row_i),(-1,row_i),GRAY))
            cap_total += item.precio_total; row_i += 1

        table_data.append(["","",f"SUBTOTAL {cap}","","","",f"${cap_total:,.0f}",""])
        styles += [("BACKGROUND",(0,row_i),(-1,row_i),LIGHT),
                   ("FONTNAME",(0,row_i),(-1,row_i),"Helvetica-Bold"),
                   ("TEXTCOLOR",(0,row_i),(-1,row_i),STEEL)]
        row_i += 1

    for lbl, val in [
        ("SUBTOTAL DIRECTO", p.resumen.subtotal_directo),
        (f"AIU {p.resumen.porcentaje_aiu}%", p.resumen.aiu),
        ("BASE CON AIU", p.resumen.subtotal_con_aiu),
        (f"IVA {p.resumen.porcentaje_iva:.0f}%", p.resumen.iva),
        ("TOTAL PRESUPUESTO", p.resumen.total),
    ]:
        is_t = lbl == "TOTAL PRESUPUESTO"
        bg = NAVY if is_t else colors.HexColor("#D0E4F7")
        fg = colors.white if is_t else NAVY
        table_data.append(["","",lbl,"","","",f"${val:,.0f}",""])
        styles += [("BACKGROUND",(0,row_i),(-1,row_i),bg),
                   ("TEXTCOLOR",(0,row_i),(-1,row_i),fg),
                   ("FONTNAME",(0,row_i),(-1,row_i),"Helvetica-Bold"),
                   ("FONTSIZE",(0,row_i),(-1,row_i),9 if is_t else 7)]
        row_i += 1

    story.append(Table(table_data, colWidths=COL_W, repeatRows=1,
                       style=TableStyle(styles)))
    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=BORDER))
    story.append(Spacer(1, 0.2*cm))
    for nota in p.notas:
        story.append(Paragraph(f"• {nota}", s_note))

    doc.build(story); buf.seek(0)
    return buf.getvalue()
