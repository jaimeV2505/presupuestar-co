# -*- coding: utf-8 -*-
"""
TEST DE VIAJE PUBLICO — el segundo robot: la historia de una OBRA PUBLICA.

El ingeniero gana un contrato estatal y lo administra completo DENTRO de la
plataforma (sin cliente final): crea el proyecto publico -> construye sus
propios APUs (manual, duplicado de la base 2026, y COMPUESTO con precios de
SU proveedor) -> presupuesto -> candado anti-compartir -> avance -> acta
parcial (cuenta) -> adicional aprobado internamente -> termina y liquida
la retegarantia. Todo con los candados de la casa.

Correr local: cd backend && python3 tests/test_viaje_publico.py
"""
import os
import sys
import tempfile

os.environ.setdefault("DATABASE_URL", f"sqlite:///{tempfile.mkdtemp()}/publico.db")
os.environ.setdefault("JWT_SECRET", "secreto-viaje-publico-0123456789abcdef")  # gitleaks:allow (sandbox del mundo efimero)

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from fastapi.testclient import TestClient  # noqa: E402
from app.main import app                   # noqa: E402

c = TestClient(app)
PASOS = []


def paso(nombre, resp, status=200, contiene=None):
    ok = resp.status_code == status
    cuerpo = resp.json() if resp.headers.get("content-type", "").startswith("application/json") else {}
    if ok and contiene:
        ok = all(k in str(cuerpo) for k in contiene)
    print(f"  {'✓' if ok else '❌'} {nombre}")
    if not ok:
        print(f"     status={resp.status_code} esperado={status}")
        print(f"     cuerpo={str(cuerpo)[:400]}")
        raise SystemExit(f"VIAJE PUBLICO ROTO EN: {nombre}")
    PASOS.append(nombre)
    return cuerpo


print("═══ ACTO 0: EL INGENIERO Y SU CONTRATO ESTATAL ═══")
d = paso("registro", c.post("/api/auth/registro",
         json={"nombre": "Ing Publico", "email": "ing@publico.test",
               "password": "ObraPublica2026", "ciudad": "bogota"}))
H = {"Authorization": f"Bearer {d['token']}"}

d = paso("crear proyecto PUBLICO (entidad + contrato)", c.post("/api/proyectos", json={
    "nombre": "Mantenimiento via terciaria vereda El Roble",
    "sector": "publico",
    "entidad_nombre": "Alcaldia de Prueba - Secretaria de Infraestructura",
    "contrato_numero": "OP-2026-0457",
    "supervisor_nombre": "Ing. Supervisora Perez",
    "cliente_nombre": "", "region": "bogota"}, headers=H))
PID = d["id"]
d = paso("detalle refleja el sector", c.get(f"/api/proyectos/{PID}", headers=H),
         contiene=["publico", "OP-2026-0457"])
TOKEN = d["share_token"]

print("═══ ACTO 1: MIS APUs — las tres vias ═══")
d = paso("APU manual", c.post("/api/apus", json={
    "codigo": "MI-001", "descripcion": "Localizacion y replanteo via terciaria",
    "unidad": "m2", "precio": 3500, "sector_tag": "publico"}, headers=H))
APU_MANUAL = d["id"]

r = c.post("/api/apus", json={"descripcion": "x", "precio": 100}, headers=H)
paso("CANDADO: descripcion de 1 caracter -> 400", r, status=400)
r = c.post("/api/apus", json={"descripcion": "precio absurdo", "precio": 3_000_000_000}, headers=H)
paso("CANDADO: precio fuera de rango -> 400", r, status=400)

# duplicar de la base 2026 (la base JAMAS se toca)
import json as _json
base = _json.load(open(os.path.join(os.path.dirname(__file__), "..", "app", "data", "apu_2026.json")))
items_base = base if isinstance(base, list) else list(base.values())[0]
CODIGO_BASE = items_base[0]["codigo"]
d = paso("duplicar de la base como MIO", c.post(
    f"/api/apus/duplicar-de-base?codigo={CODIGO_BASE}&region=bogota", headers=H),
    contiene=["MI-"])
assert d["origen_base"] == CODIGO_BASE
d = paso("EDITAR mi APU (el bug del doble body, vigilado)", c.put(f"/api/apus/{APU_MANUAL}", json={
    "codigo": "MI-001", "descripcion": "Localizacion y replanteo via terciaria (rev.1)",
    "unidad": "m2", "precio": 3600, "sector_tag": "publico"}, headers=H))
assert d["precio"] == 3600 and "rev.1" in d["descripcion"]

r = c.post("/api/apus/duplicar-de-base?codigo=NO-EXISTE-999", headers=H)
paso("CANDADO: codigo inexistente en la base -> 404", r, status=404)

print("═══ ACTO 1.5: EL CONSTRUCTOR (proveedor propio + composicion) ═══")
d = paso("crear MI proveedor", c.post("/api/proveedores", json={
    "nombre": "Ferreteria El Roble", "ciudad": "Bogota", "telefono": "3001112233"}, headers=H))
PROV = d["id"]
paso("capturar precio del cemento", c.post(f"/api/proveedores/{PROV}/precios",
     json={"insumo": "Cemento gris", "unidad": "bulto", "precio": 28500}, headers=H))
r = c.post(f"/api/proveedores/{PROV}/precios",
           json={"insumo": "x", "precio": -5}, headers=H)
paso("CANDADO: precio negativo -> 400", r, status=400)

d = paso("buscar insumo MULTI-FUENTE (mi proveedor + referencia)",
         c.get("/api/insumos/buscar?q=cemento", headers=H))
fuentes = {r["fuente"] for r in d["resultados"]}
assert "mi_proveedor" in fuentes and "referencia" in fuentes, fuentes
mio = next(r for r in d["resultados"] if r["fuente"] == "mi_proveedor")
assert mio["precio"] == 28500 and mio["detalle"] == "Ferreteria El Roble" and mio["fecha"]

d = paso("catalogo curado: categorias", c.get("/api/insumos/catalogo", headers=H))
assert d["total"] >= 200 and len(d["categorias"]) >= 15, d
cat0 = d["categorias"][0]["nombre"]
d = paso(f"catalogo: insumos de una categoria", c.get(f"/api/insumos/catalogo?categoria={cat0}", headers=H))
assert d["insumos"] and all(i["precio"] > 0 for i in d["insumos"])
print(f"  ✓ catalogo curado navegable: 200+ insumos de referencia adentro de la plataforma")
PASOS.append("catalogo curado 200+")

d = paso("comparador: mi precio vs referencia", c.get("/api/insumos/comparar?q=cemento", headers=H))
assert d["mi_mejor"]["precio"] == 28500, d["mi_mejor"]
_ref = (d.get("referencia") or {}).get("precio")
assert _ref and _ref > 0, f"el comparador debe traer la referencia del catalogo: {d}"
# la matematica del ahorro cuadra con los valores REALES devueltos
assert d["ahorro"] == max(0, _ref - 28500), f"ahorro incoherente: ref={_ref} ahorro={d['ahorro']}"
r = c.get("/api/insumos/buscar?q=x", headers=H)
paso("CANDADO: busqueda de 1 letra -> 4xx", r, status=422) if r.status_code == 422 else paso("CANDADO: busqueda corta rechazada", r, status=400)

d = paso("componer APU con mis insumos (vista previa)", c.post("/api/apus/componer", json={
    "insumos": [
        {"nombre": "Cemento gris", "unidad": "bulto", "cantidad": 0.35, "precio": 28500, "desperdicio_pct": 5},
        {"nombre": "Arena lavada", "unidad": "m3", "cantidad": 0.04, "precio": 65000}],
    "mano_obra": 9000, "herramienta_pct": 5}, headers=H))
assert d["precio_unitario"] == 22524, d   # el mismo caso exacto del smoke
d2 = paso("crear APU y guardar su desglose compuesto", c.post("/api/apus", json={
    "codigo": "MI-COMP-1", "descripcion": "Panete 1:4 muros (compuesto)",
    "unidad": "m2", "precio": 0, "sector_tag": "publico"}, headers=H))
APU_COMP = d2["id"]
d = paso("fijar desglose", c.put(f"/api/apus/{APU_COMP}/desglose", json={
    "insumos": [
        {"nombre": "Cemento gris", "unidad": "bulto", "cantidad": 0.35, "precio": 28500, "desperdicio_pct": 5},
        {"nombre": "Arena lavada", "unidad": "m3", "cantidad": 0.04, "precio": 65000}],
    "mano_obra": 9000, "herramienta_pct": 5}, headers=H))
assert d["precio"] == 22524 and d["desglose"]["materiales"] == 13074
# EL ANALISIS ES EDITABLE: jugar con los valores -> el servidor recalcula (transporte incluido)
d = paso("editar el analisis: +$1.500 de transporte por m2", c.put(f"/api/apus/{APU_COMP}/desglose", json={
    "insumos": [
        {"nombre": "Cemento gris", "unidad": "bulto", "cantidad": 0.35, "precio": 28500, "desperdicio_pct": 5},
        {"nombre": "Arena lavada", "unidad": "m3", "cantidad": 0.04, "precio": 65000}],
    "mano_obra": 9000, "herramienta_pct": 5, "transporte": 1500}, headers=H))
assert d["precio"] == 24024 and d["desglose"]["transporte"] == 1500, d
d = paso("volver el analisis al canonico (transporte 0)", c.put(f"/api/apus/{APU_COMP}/desglose", json={
    "insumos": [
        {"nombre": "Cemento gris", "unidad": "bulto", "cantidad": 0.35, "precio": 28500, "desperdicio_pct": 5},
        {"nombre": "Arena lavada", "unidad": "m3", "cantidad": 0.04, "precio": 65000}],
    "mano_obra": 9000, "herramienta_pct": 5}, headers=H))
assert d["precio"] == 22524 and (d["desglose"].get("transporte") or 0) == 0
PASOS.append("analisis editable: recalculo en servidor con transporte (24.024 exacto)")
d = paso("mis APUs listados (filtro sector publico)", c.get("/api/apus?sector=publico", headers=H))
assert len(d["items"]) >= 3

# R1: EL RECETARIO SEMILLA — 16 recetas listas, precio del servidor, idempotente
d = paso("cargar el recetario de arranque", c.post("/api/apus/recetario", headers=H))
assert d["creados"] >= 15, d
assert "REC-PANETE" in d["codigos"]
d2 = paso("recetario idempotente (2a vez: cero duplicados)", c.post("/api/apus/recetario", headers=H))
assert d2["creados"] == 0 and d2["ya_existian"] >= 15, d2
d = paso("las recetas viven en Mis APUs con su analisis", c.get("/api/apus", headers=H))
panete = next(a for a in d["items"] if a["codigo"] == "REC-PANETE")
assert panete["precio"] == 22524, f"la receta semilla del panete debia dar 22.524: {panete['precio']}"
assert panete["desglose"] and panete["desglose"]["insumos"], "receta semilla sin desglose"
PASOS.append("recetario semilla: 16 recetas, panete 22.524 exacto, idempotente")

print("═══ ACTO 1.9: EL LADRON (ownership cruzado) ═══")
d2 = paso("registro del intruso", c.post("/api/auth/registro",
          json={"email": "intruso@publico.test", "password": "Intruso2026x"}))
H2 = {"Authorization": f"Bearer {d2['token']}"}
r = c.put(f"/api/apus/{APU_MANUAL}", json={"descripcion": "hackeado", "precio": 1}, headers=H2)
paso("CANDADO: editar APU ajeno -> 404", r, status=404)
r = c.delete(f"/api/apus/{APU_COMP}", headers=H2)
paso("CANDADO: borrar APU ajeno -> 404", r, status=404)
r = c.get(f"/api/proveedores/{PROV}/precios", headers=H2)
paso("CANDADO: espiar precios de proveedor ajeno -> 404", r, status=404)
d = c.get("/api/apus", headers=H2).json()
assert d["total"] == 0, "el intruso VE apus ajenos!"
print("  ✓ el intruso no ve, no edita, no borra nada ajeno")
PASOS.append("ownership cruzado blindado")

print("═══ ACTO 2: EL PRESUPUESTO OFICIAL Y LOS CANDADOS DEL SECTOR ═══")
paso("items desde MIS APUs + contrato (anticipo 30%, rete 10%)", c.put(f"/api/proyectos/{PID}", json={
    "items": [
        {"id": "it1", "descripcion": "Localizacion y replanteo", "unidad": "m2",
         "cantidad": 1200, "precio_unitario": 3500},
        {"id": "it2", "descripcion": "Panete 1:4 muros (compuesto)", "unidad": "m2",
         "cantidad": 300, "precio_unitario": 22524, "apu_id": APU_COMP},
    ],
    "aiu": {"admin": 20, "imprevistos": 3, "utilidad": 7, "aplicar": True, "iva_sobre_utilidad": True},
    "contrato": {"plazo_dias": 90, "anticipo_pct": 30, "retegarantia_pct": 10,
                 "fecha_inicio": "2026-09-01", "lugar": "Vereda El Roble",
                 "deducciones": [
                     {"nombre": "Contribución obra pública 5% (Ley 1106/2006)", "pct": 5},
                     {"nombre": "Estampillas territoriales", "pct": 4},
                     {"nombre": "Retefuente construcción", "pct": 2}]},
}, headers=H))

r = c.put(f"/api/proyectos/{PID}", json={"contrato": {"anticipo_pct": 30, "retegarantia_pct": 10,
    "deducciones": [{"nombre": "Estampilla absurda", "pct": 30}]}}, headers=H)
paso("CANDADO: deduccion del 30% -> 400", r, status=400)
# restaurar el contrato bueno (el candado no debe haberlo tocado)
paso("contrato restaurado con las 3 de ley", c.put(f"/api/proyectos/{PID}", json={
    "contrato": {"plazo_dias": 90, "anticipo_pct": 30, "retegarantia_pct": 10,
                 "fecha_inicio": "2026-09-01", "lugar": "Vereda El Roble",
                 "deducciones": [
                     {"nombre": "Contribución obra pública 5% (Ley 1106/2006)", "pct": 5},
                     {"nombre": "Estampillas territoriales", "pct": 4},
                     {"nombre": "Retefuente construcción", "pct": 2}]}}, headers=H))

r = c.post(f"/api/share/proyectos/{PID}/compartir", headers=H)
paso("CANDADO ESTRELLA: obra publica NO comparte -> 400", r, status=400,
     ) if r.status_code == 400 else paso("candado publico", r, status=400)

# 🔒 DOCTRINA: el adicional exige el SELLO — antes, el presupuesto se edita directo
r = c.post("/api/otrosies", json={
    "proyecto_id": PID, "motivo": "Ajuste previo al inicio",
    "items": [{"id": "pre1", "descripcion": "Señalizacion preventiva", "unidad": "un",
               "cantidad": 10, "precio_unitario": 45000}]}, headers=H)
paso("CANDADO: adicional ANTES del sello -> 400 (edita el presupuesto directo)", r, status=400)

print("═══ ACTO 3: EJECUCION — EL PRIMER AVANCE SELLA EL CONTRATO ═══")
d = paso("proyecto sigue en borrador", c.get(f"/api/proyectos/{PID}", headers=H))
assert d["estado"] == "borrador", d["estado"]
d = paso("avance 50%", c.post(f"/api/avances/proyectos/{PID}/avances", json={
    "titulo": "Acta parcial No. 1", "items": [{"id": "it1", "pct": 100}, {"id": "it2", "pct": 30}],
}, headers=H))
AV1 = d["id"]
d = paso("el primer avance sello el contrato (aceptado automatico)",
         c.get(f"/api/proyectos/{PID}", headers=H))
assert d["estado"] == "aceptado", f"el avance no sello: {d['estado']}"
r = c.put(f"/api/proyectos/{PID}", json={"items": [
    {"id": "hack", "descripcion": "cambiar presupuesto oficial", "unidad": "un",
     "cantidad": 1, "precio_unitario": 1}]}, headers=H)
paso("CANDADO: presupuesto oficial sellado (items -> 400)", r, status=400)

# Ahora SI: el adicional post-sello (la via legitima de los cambios)
d_pre = paso("adicional post-sello (la via correcta)", c.post("/api/otrosies", json={
    "proyecto_id": PID, "motivo": "Ajuste previo al inicio",
    "items": [{"id": "pre1", "descripcion": "Señalizacion preventiva", "unidad": "un",
               "cantidad": 10, "precio_unitario": 45000}]}, headers=H))
paso("y se aprueba internamente", c.post(f"/api/otrosies/{d_pre['id']}/aprobar-interno", headers=H))
r = c.post("/api/exportar/apus-pdf", json={"proyecto_id": PID}, headers=H)
paso("ANEXO DE PROPUESTA: APUs detallados en PDF", r)
assert r.content[:4] == b"%PDF", "el anexo de APUs no es un PDF"
assert len(r.content) > 2000, "el anexo de APUs salio vacio"
PASOS.append("anexo de APUs detallados (formato propuesta)")

# F2: la LISTA DE COMPRAS de la obra — consolidada, con desperdicio y MI precio
d = paso("explosion de insumos (lista de materiales)",
         c.post("/api/exportar/explosion", json={"proyecto_id": PID}, headers=H))
_ins = {x["nombre"]: x for x in d["insumos"]}
# panete 300 m2 × 0.35 bulto × 1.05 desperdicio = 110.25 bultos exactos
assert _ins["Cemento gris"]["cantidad"] == 110.25, _ins["Cemento gris"]
assert _ins["Arena lavada"]["cantidad"] == 12.0, _ins["Arena lavada"]
# cruzado con MI proveedor (Ferreteria El Roble a $28.500, con fecha)
assert _ins["Cemento gris"]["fuente_precio"] == "mi_proveedor"
assert _ins["Cemento gris"]["proveedor"] == "Ferreteria El Roble" and _ins["Cemento gris"]["fecha_precio"]
assert _ins["Cemento gris"]["subtotal"] == round(110.25 * 28500), _ins["Cemento gris"]["subtotal"]
# el parte honesto: it1 (replanteo, sin desglose) queda declarado
assert d["items_sin_desglose"] >= 1 and d["items_explotados"] >= 1
r = c.post("/api/exportar/explosion-excel", json={"proyecto_id": PID}, headers=H)
paso("lista de materiales en Excel (para la ferreteria)", r)
assert r.content[:2] == b"PK", "el Excel de materiales no es un xlsx"
PASOS.append("explosion de insumos exacta + cruce con proveedor + Excel")

# EL FORMATO DEL OFICIO: el Excel principal trae las 3 hojas del formato real
r = c.post("/api/exportar/excel", json={"proyecto_id": PID}, headers=H)
paso("Excel del presupuesto en formato del oficio (3 hojas)", r)
from openpyxl import load_workbook as _lwb
import io as _io
_wb = _lwb(_io.BytesIO(r.content))
assert _wb.sheetnames == ["Presupuesto", "Memoria de cantidades", "Materiales"], _wb.sheetnames
_txt = " ".join(str(cel.value) for fila in _wb["Presupuesto"].iter_rows() for cel in fila if cel.value)
assert "VALOR TOTAL OBRA" in _txt and "ELABORÓ" in _txt, "faltan las filas de cierre del oficio"
assert "1.1" in _txt, "falta la numeracion jerarquica de items"
_mat = " ".join(str(cel.value) for fila in _wb["Materiales"].iter_rows() for cel in fila if cel.value)
assert "Cemento gris" in _mat and "TOTAL MATERIALES" in _mat, "la hoja Materiales no trae la explosion viva"
_mem = " ".join(str(cel.value) for fila in _wb["Memoria de cantidades"].iter_rows() for cel in fila if cel.value)
assert "MEMORIA DE CANTIDADES" in _mem and "Panete" in _mem, "la memoria no lista las actividades"
PASOS.append("Excel formato del oficio: 3 hojas, numeracion 1.1, VALOR TOTAL OBRA, ELABORO, memoria y materiales vivos")
r = c.put(f"/api/proyectos/{PID}", json={"estado": "borrador"}, headers=H)
paso("CANDADO: degradar el estado tras el sello -> 400", r, status=400)
r = c.put(f"/api/proyectos/{PID}", json={"estado": "entrega_solicitada"}, headers=H)
paso("CANDADO: estados del enlace de cliente en publico -> 400", r, status=400)
print("  ✓ semantica publica: herramientas libres, el primer avance sella, cambios via adicionales")
PASOS.append("primer avance sella el presupuesto oficial")

d = paso("acta parcial (cuenta) liquidada con anticipo 30% y rete 10%",
         c.post(f"/api/cuentas/proyectos/{PID}/cuentas", json={"avance_id": AV1}, headers=H))
CID = d.get("id") or d.get("cuenta", {}).get("id")
cc = d.get("cuenta") or d
CORTE = cc["valor_corte"]; RETE = cc["retencion"]; AMORT = cc["amortizacion"]
assert RETE == round(CORTE * 0.10), f"rete no es 10%: {RETE} de {CORTE} (BUG #2 vigilado)"
DED = cc["deducciones"]
esperado_ded = round(CORTE*0.05) + round(CORTE*0.04) + round(CORTE*0.02)
assert DED == esperado_ded, f"deducciones {DED} != {esperado_ded}"
assert len(cc["deducciones_detalle"]) == 3 and "Ley 1106" in cc["deducciones_detalle"][0]["nombre"]
assert cc["neto"] == CORTE - AMORT - RETE - DED, "el NETO REAL no cuadra"
print(f"  ✓ acta publica liquidada a peso: corte {CORTE:,} - amort {AMORT:,} - rete {RETE:,} - deducciones {DED:,} = neto {cc['neto']:,}")
PASOS.append("deducciones de ley exactas")
paso("acta cobrada (pagada por la entidad)", c.post(f"/api/cuentas/{CID}/pagada", headers=H))

print("═══ ACTO 3.5: ADICIONAL / MAYORES CANTIDADES (aprobacion interna) ═══")
d = paso("proponer adicional", c.post("/api/otrosies", json={
    "proyecto_id": PID, "motivo": "Mayores cantidades por inestabilidad del terreno",
    "items": [{"id": "ex1", "descripcion": "Recebo compactado adicional", "unidad": "m3",
               "cantidad": 80, "precio_unitario": 95000}]}, headers=H))
OID = d["id"]
paso("aprobar INTERNAMENTE (sin cliente — es obra publica)",
     c.post(f"/api/otrosies/{OID}/aprobar-interno", headers=H))
r = c.post(f"/api/otrosies/{OID}/aprobar-interno", headers=H)
paso("CANDADO: doble aprobacion -> 400", r, status=400)

paso("avance final con AMBOS adicionales (N.1 previo + N.2 recebo)",
     c.post(f"/api/avances/proyectos/{PID}/avances", json={
    "titulo": "Acta parcial No. 2 — final",
    "items": [{"id": "it1", "pct": 100}, {"id": "it2", "pct": 100},
              {"id": "ot1:pre1", "pct": 100}, {"id": "ot2:ex1", "pct": 100}],
}, headers=H))

print("═══ ACTO 3.75: BALANCE DE OBRA (oficial vs ejecutado) ═══")
d = paso("balance del proyecto", c.get(f"/api/proyectos/{PID}/balance", headers=H))
rs = d["resumen"]
VALOR_BASE = 1200*3500 + 300*22524          # 10.957.200
VALOR_ADIC = 80*95000 + 10*45000             # 7.600.000 + 450.000 (el previo)
assert rs["valor_contrato_base"] == VALOR_BASE, rs
assert rs["valor_adicionales"] == VALOR_ADIC, rs
assert rs["valor_total"] == VALOR_BASE + VALOR_ADIC
assert rs["avance_fisico_pct"] == 100.0
assert rs["facturado"] > 0 and rs["pagado"] > 0
adicionales = [f for f in d["filas"] if f["es_adicional"]]
assert {f["id"] for f in adicionales} == {"ot1:pre1", "ot2:ex1"}, adicionales
print("  ✓ el balance cuadra a peso: base + adicional + avances + actas")
PASOS.append("balance de obra exacto")

print("═══ ACTO 3.9: NADA VIAJA FUERA (la doctrina del publico) ═══")
# 1) el enlace de cliente sigue bloqueado (ya probado en ACTO 2) y ademas:
# 2) un token LEGADO (proyecto que fue privado, se compartio, y luego paso a publico) no sirve
d2p = paso("proyecto privado auxiliar", c.post("/api/proyectos", json={
    "nombre": "Legado privado", "cliente_nombre": "Cliente X", "region": "bogota"}, headers=H))
PID2 = d2p["id"]
d2p = paso("compartirlo (aun privado)", c.post(f"/api/share/proyectos/{PID2}/compartir", headers=H))
TOKEN_LEGADO = d2p["token"]
paso("cambiarlo a publico (antes de firma se puede)", c.put(f"/api/proyectos/{PID2}",
     json={"sector": "publico", "entidad_nombre": "Entidad Y"}, headers=H))
r = c.get(f"/api/share/publico/{TOKEN_LEGADO}")
paso("CANDADO: el token legado YA NO sirve -> 404", r, status=404)
# La doctrina cubre TODAS las puertas, no solo la ventana (los 8 POST del token)
r = c.post(f"/api/share/publico/{TOKEN_LEGADO}/aceptar",
           json={"nombre": "Intruso Externo", "documento": "0", "firma_imagen": ""})
paso("CANDADO: FIRMAR con token legado -> 404 (nadie sella desde afuera)", r, status=404)
r = c.post(f"/api/share/publico/{TOKEN_LEGADO}/reportar-pendientes",
           json={"detalle": "intento de intrusion por token viejo"})
paso("CANDADO: reportar pendientes con token legado -> 404", r, status=404)
d = paso("el detalle del proyecto convertido ya NO expone token", c.get(f"/api/proyectos/{PID2}", headers=H))
assert not d.get("share_token"), "al convertir a publico, share_token debia quedar limpio"
print("  ✓ doctrina sellada: en obra publica, ni tokens viejos abren puertas — GET ni POST")
PASOS.append("tokens legados de publicos rechazados (GET + POST + token limpio)")

print("═══ ACTO 4: LIQUIDACION DEL CONTRATO ═══")
# Duplicar un contrato publico conserva sector, entidad y condiciones (bug R3 cazado)
d = paso("duplicar el contrato publico", c.post(f"/api/proyectos/{PID}/duplicar", headers=H))
assert d.get("sector") == "publico", "la copia perdio el sector publico"
assert d.get("contrato_numero") == "OP-2026-0457", "la copia perdio el numero de contrato"
PID_COPIA = d["id"]
paso("borrar la copia (limpieza)", c.delete(f"/api/proyectos/{PID_COPIA}", headers=H))

paso("terminar y liquidar (endpoint dedicado, un clic)",
     c.post(f"/api/proyectos/{PID}/terminar", headers=H))
r = c.post(f"/api/proyectos/{PID}/terminar", headers=H)
paso("CANDADO: terminar dos veces -> 400", r, status=400)
r = c.put(f"/api/proyectos/{PID}", json={"estado": "aceptado"}, headers=H)
paso("CANDADO: reabrir un contrato terminado -> 400", r, status=400)
# 🔒 Terminado = SOLO LECTURA total: ni nombre, ni avances, ni gastos
r = c.put(f"/api/proyectos/{PID}", json={"nombre": "renombrar terminado"}, headers=H)
paso("CANDADO: editar CUALQUIER campo de un terminado -> 400 (solo lectura)", r, status=400)
r = c.post(f"/api/avances/proyectos/{PID}/avances",
           json={"titulo": "avance postumo", "items": [{"id": "it1", "pct": 100}]}, headers=H)
paso("CANDADO: avance sobre terminado -> 400", r, status=400)
r = c.post(f"/api/gastos/proyectos/{PID}/gastos",
           json={"descripcion": "gasto postumo", "valor": 1000}, headers=H)
paso("CANDADO: gasto sobre terminado -> 400", r, status=400)
# 🎨 DOCTRINA: los disenos son del universo PRIVADO — en publico no existen
r = c.post(f"/api/disenos/proyectos/{PID}/disenos",
           json={"titulo": "render", "imagen_b64": "data:image/jpeg;base64,AAAA"}, headers=H)
paso("CANDADO: disenos en obra publica -> 400 (nada viaja fuera)", r, status=400)
d = paso("duplicar un TERMINADO si esta permitido (nace borrador editable, con nombre propio)",
         c.post(f"/api/proyectos/{PID}/duplicar", json={"nombre": "Via El Roble FASE 2"}, headers=H))
assert d.get("estado") == "borrador" and d.get("sector") == "publico", d
assert d.get("nombre") == "Via El Roble FASE 2", "el nombre propio de la copia no se respeto"
paso("la copia SI se edita", c.put(f"/api/proyectos/{d['id']}", json={"nombre": "Copia editable"}, headers=H))
# INDEPENDENCIA: editar la copia jamas toca al original (cero herencia post-creacion)
paso("items nuevos en la copia", c.put(f"/api/proyectos/{d['id']}", json={"items": [
    {"id": "solo-copia", "descripcion": "Item exclusivo de la copia", "unidad": "un",
     "cantidad": 1, "precio_unitario": 999}]}, headers=H))
d_orig = paso("el ORIGINAL sigue intacto", c.get(f"/api/proyectos/{PID}", headers=H))
assert d_orig["estado"] == "terminado", "editar la copia cambio el estado del original!"
assert all(i["id"] != "solo-copia" for i in d_orig.get("items", [])), "la copia heredo items al original!"
assert d_orig["nombre"].startswith("Mantenimiento via"), "el nombre del original cambio!"
paso("borrar la copia del terminado", c.delete(f"/api/proyectos/{d['id']}", headers=H))
PASOS.append("terminado = solo lectura (PUT/avance/gasto -> 400) + duplicar con nombre + copias independientes")
d = paso("liberar retegarantia", c.post(f"/api/cuentas/proyectos/{PID}/retegarantia", headers=H))
r = c.post(f"/api/cuentas/proyectos/{PID}/retegarantia", headers=H)
paso("CANDADO: liberar dos veces -> 400", r, status=400)

paso("borrar el proyecto auxiliar (cascada)", c.delete(f"/api/proyectos/{PID2}", headers=H))
assert c.get(f"/api/proyectos/{PID2}", headers=H).status_code == 404

print(f"\n{'='*52}")
print(f"VIAJE PUBLICO: {len(PASOS)} pasos en verde ✅")
print("El ingeniero administro su contrato estatal completo — sin salir de la plataforma.")
